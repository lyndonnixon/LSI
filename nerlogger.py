# -*- coding: UTF-8 -*-
"""Logger that can save the invocation results of Zemanta API and NERD API as a Excel spreadsheet.

.. moduleauthor:: Dong Liu <liu.dong66@gmail.com>
"""
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter


class NERLogger(object):
    """Save NER service invocation results to a Excel spreadsheet"""
    def __init__(self):
        self.ner_log_workbook = Workbook()
        self.ner_log_worksheet = self.ner_log_workbook.worksheets[0]
        self.ner_log_worksheet.title = 'NER service invocation results'
        self.ner_log_worksheet.cell('%s%s' % (get_column_letter(1), 1)).value = "Title"
        self.ner_log_worksheet.cell('%s%s' % (get_column_letter(2), 1)).value = "Description"
        self.ner_log_worksheet.cell('%s%s' % (get_column_letter(3), 1)).value = "Zemanta Results"
        self.ner_log_worksheet.cell('%s%s' % (get_column_letter(4), 1)).value = "NERD Results"
        self.cur_row = 2

    def log(self, title, description, zemanta_results, nerd_results):
        """Create a log item.

        Args:
            title (str): title of media resource.

            description (str): description of media resource.

            zemanta_results (str): results of Zemanta API invocation.

            nerd_results (str): results of NERD API invocation.
        """
        self.ner_log_worksheet.cell('%s%s' % (get_column_letter(1), self.cur_row)).value = title
        self.ner_log_worksheet.cell('%s%s' % (get_column_letter(2), self.cur_row)).value = description
        self.ner_log_worksheet.cell('%s%s' % (get_column_letter(3), self.cur_row)).value = zemanta_results
        self.ner_log_worksheet.cell('%s%s' % (get_column_letter(4), self.cur_row)).value = nerd_results
        self.cur_row += 1

    def save(self, xls_file):
        """Save log items as Excel files.

        Args:
            xls_file (str): path to the Excel file.
        """
        ew = ExcelWriter(workbook=self.ner_log_workbook)
        ew.save(filename=xls_file)
